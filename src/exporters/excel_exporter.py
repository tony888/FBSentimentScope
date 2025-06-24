"""Excel exporter for analysis results."""

from typing import List
import logging
from datetime import datetime

try:
    import pandas as pd
    import xlsxwriter
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError as e:
    EXCEL_AVAILABLE = False
    EXCEL_IMPORT_ERROR = str(e)

from .base_exporter import BaseExporter
from ..core.models import AnalysisResult
from ..core.exceptions import ExportError

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """Export analysis results to Excel format with multiple sheets and formatting."""
    
    def __init__(self, output_dir: str = "exports"):
        """Initialize the Excel exporter.
        
        Args:
            output_dir: Directory to save exported files
            
        Raises:
            ExportError: If required Excel libraries are not available
        """
        super().__init__(output_dir)
        
        if not EXCEL_AVAILABLE:
            raise ExportError(f"Excel export requires pandas, xlsxwriter, and openpyxl: {EXCEL_IMPORT_ERROR}")
    
    def export(self, results: List[AnalysisResult], filename: str, **kwargs) -> str:
        """Export analysis results to Excel file with multiple sheets.
        
        Args:
            results: List of analysis results to export
            filename: Name of the output file (without extension)
            **kwargs: Additional export options
                include_charts: Whether to include charts (default: True)
                
        Returns:
            Path to the exported Excel file
            
        Raises:
            ExportError: If export fails
        """
        include_charts = kwargs.get('include_charts', True)
        output_file = self.output_dir / f"{filename}.xlsx"
        
        try:
            # Prepare data
            data = self._prepare_data(results)
            summary_stats = self._get_summary_stats(results)
            
            if not data:
                logger.warning("No data to export")
                raise ExportError("No data to export")
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create Excel workbook with multiple sheets
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Main data sheet
                df.to_excel(writer, sheet_name='Analysis Results', index=False)
                
                # Summary sheet
                summary_df = pd.DataFrame(list(summary_stats.items()), columns=['Metric', 'Value'])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Separate sheets for posts and comments
                posts_df = df[df['type'] == 'post'].copy()
                comments_df = df[df['type'] == 'comment'].copy()
                
                if not posts_df.empty:
                    posts_df.to_excel(writer, sheet_name='Posts', index=False)
                
                if not comments_df.empty:
                    comments_df.to_excel(writer, sheet_name='Comments', index=False)
                
                # Format the workbook
                self._format_workbook(writer)
                
                # Add charts if requested
                if include_charts:
                    self._add_charts(writer, df, summary_stats)
            
            logger.info(f"Successfully exported {len(data)} rows to {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}")
            raise ExportError(f"Failed to export to Excel: {e}")
    
    def _format_workbook(self, writer) -> None:
        """Apply formatting to the Excel workbook.
        
        Args:
            writer: ExcelWriter object
        """
        try:
            workbook = writer.book
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Format headers
                for cell in worksheet[1]:  # First row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Failed to format workbook: {e}")
    
    def _add_charts(self, writer, df: pd.DataFrame, summary_stats: dict) -> None:
        """Add charts to the Excel workbook.
        
        Args:
            writer: ExcelWriter object
            df: Main DataFrame with analysis results
            summary_stats: Summary statistics dictionary
        """
        try:
            workbook = writer.book
            
            # Create a charts sheet
            charts_sheet = workbook.create_sheet("Charts")
            
            # Add sentiment distribution data
            sentiment_data = [
                ['Sentiment', 'Average Score'],
                ['Positive', summary_stats.get('avg_positive', 0)],
                ['Negative', summary_stats.get('avg_negative', 0)],
                ['Neutral', summary_stats.get('avg_neutral', 0)],
            ]
            
            for row_idx, row in enumerate(sentiment_data, 1):
                for col_idx, value in enumerate(row, 1):
                    charts_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Format the charts sheet header
            for cell in charts_sheet[1]:
                cell.font = Font(bold=True)
            
            logger.info("Charts added to Excel workbook")
            
        except Exception as e:
            logger.warning(f"Failed to add charts: {e}")
    
    def export_dashboard(self, results: List[AnalysisResult], filename: str) -> str:
        """Export a dashboard-style Excel file with enhanced formatting and charts.
        
        Args:
            results: List of analysis results to export
            filename: Name of the output file (without extension)
            
        Returns:
            Path to the exported Excel file
            
        Raises:
            ExportError: If export fails
        """
        output_file = self.output_dir / f"{filename}_dashboard.xlsx"
        
        try:
            # Prepare comprehensive data
            data = self._prepare_data(results)
            summary_stats = self._get_summary_stats(results)
            
            if not data:
                raise ExportError("No data to export")
            
            df = pd.DataFrame(data)
            
            # Create comprehensive dashboard
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                workbook = writer.book
                
                # Dashboard overview sheet
                dashboard_sheet = workbook.add_worksheet('Dashboard')
                
                # Add title
                title_format = workbook.add_format({
                    'bold': True,
                    'font_size': 16,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#366092',
                    'font_color': 'white'
                })
                
                dashboard_sheet.merge_range('A1:D1', 'Facebook Comment Analysis Dashboard', title_format)
                
                # Add summary statistics
                header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BD'})
                dashboard_sheet.write('A3', 'Summary Statistics', header_format)
                
                row = 4
                for key, value in summary_stats.items():
                    dashboard_sheet.write(row, 0, key.replace('_', ' ').title())
                    dashboard_sheet.write(row, 1, value)
                    row += 1
                
                # Add detailed data to separate sheets
                df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                # Posts and Comments sheets
                posts_df = df[df['type'] == 'post'].copy()
                comments_df = df[df['type'] == 'comment'].copy()
                
                if not posts_df.empty:
                    posts_df.to_excel(writer, sheet_name='Posts Analysis', index=False)
                
                if not comments_df.empty:
                    comments_df.to_excel(writer, sheet_name='Comments Analysis', index=False)
            
            logger.info(f"Dashboard exported to {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"Failed to export dashboard: {e}")
            raise ExportError(f"Failed to export dashboard: {e}")
